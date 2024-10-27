import itertools
import openpyxl
import numpy as np
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.statespace.sarimax import SARIMAX
from matplotlib import pyplot as plt

T = 60 * 60 * 24 * 30

area_sup = 81000 * 10 ** 6
area_eri = 25700 * 10 ** 6
area_ont = 19000 * 10 ** 6
area_hm = 118000 * 10 ** 6

deep_cyan = '#30688D'
light_green = '#35B777'


prc_sup_wb = openpyxl.load_workbook('prc_sup_lake_mon.xlsx', data_only=True)
eva_sup_wb = openpyxl.load_workbook('evaporation_sup.xlsx', data_only=True)
runoff_sup_wb = openpyxl.load_workbook('runoff_sup_arm.xlsx', data_only=True)

prc_hm_wb = openpyxl.load_workbook('prc_mic_lake_mon.xlsx', data_only=True)
eva_hm_wb = openpyxl.load_workbook('evaporation_mic.xlsx', data_only=True)
runoff_hur_wb = openpyxl.load_workbook('runoff_hur_arm.xlsx', data_only=True)
runoff_mic_wb = openpyxl.load_workbook('runoff_mic_arm.xlsx', data_only=True)

prc_eri_wb = openpyxl.load_workbook('prc_eri_lake_mon.xlsx', data_only=True)
eva_eri_wb = openpyxl.load_workbook('evaporation_eri.xlsx', data_only=True)
runoff_eri_wb = openpyxl.load_workbook('runoff_eri_arm.xlsx', data_only=True)

prc_ont_wb = openpyxl.load_workbook('prc_ont_lake_mon.xlsx', data_only=True)
eva_ont_wb = openpyxl.load_workbook('evaporation_ont.xlsx', data_only=True)
runoff_ont_wb = openpyxl.load_workbook('runoff_ont_arm.xlsx', data_only=True)

prc_sup_ws = prc_sup_wb['prc_sup_lake_mon']
eva_sup_ws = eva_sup_wb['evaporation_sup']
runoff_sup_ws = runoff_sup_wb['runoff_sup_arm']

prc_hm_ws = prc_hm_wb['prc_mic_lake_mon']
eva_hm_ws = eva_hm_wb['evaporation_mic']
runoff_hur_ws = runoff_hur_wb['runoff_hur_arm']
runoff_mic_ws = runoff_mic_wb['runoff_mic_arm']

prc_eri_ws = prc_eri_wb['prc_eri_lake_mon']
eva_eri_ws = eva_eri_wb['evaporation_eri']
runoff_eri_ws = runoff_eri_wb['runoff_eri_arm']

prc_ont_ws = prc_ont_wb['prc_ont_lake_mon']
eva_ont_ws = eva_ont_wb['evaporation_ont']
runoff_ont_ws = runoff_ont_wb['runoff_ont_arm']

prc_sup = np.array([prc_sup_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(7, 88), range(2, 14))])[:96]
eva_sup = np.array([eva_sup_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(67, 75), range(2, 14))])
runoff_sup = np.array([runoff_sup_ws.cell(row=i, column=3).value for i in range(1252, 1348)])

prc_hm = np.array([prc_hm_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(7, 88), range(2, 14))])[:96]
eva_hm = np.array([eva_hm_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(67, 75), range(2, 14))])
runoff_hur = np.array([runoff_hur_ws.cell(row=i, column=3).value for i in range(1252, 1348)])
runoff_mic = np.array([runoff_mic_ws.cell(row=i, column=3).value for i in range(1252, 1348)])

prc_eri = np.array([prc_eri_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(7, 88), range(2, 14))])[:96]
eva_eri = np.array([eva_eri_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(67, 75), range(2, 14))])
runoff_eri = np.array([runoff_eri_ws.cell(row=i, column=3).value for i in range(1252, 1348)])

prc_ont = np.array([prc_ont_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(7, 88), range(2, 14))])[:96]
eva_ont = np.array([eva_ont_ws.cell(row=i, column=j).value
                    for i,j in itertools.product(range(67, 75), range(2, 14))])
runoff_ont = np.array([runoff_ont_ws.cell(row=i, column=3).value for i in range(1252, 1348)])





def adfuller_test(data):
    result = adfuller(data)
    labels = ['ADF Test Statistic','p-value','#Lags Used','Number of Observations']
    for value, label in zip(result , labels):
        print(label + ' : ' + str(value))
    if result[1] <= 0.05:
        print("strong evidence against the null hypothesis(Ho), reject the null hypothesis. Data is stationary")
    else:
        print("weak evidence against null hypothesis,indicating it is non-stationary ")


def sarimax_predict(data):
    model = SARIMAX(data, order=(1,0,0), seasonal_order=(1,1,1,12))
    model_fit = model.fit()
    print(model_fit.summary())
    predict = model_fit.predict(start=96, end=96 + 11, dynamic=True)[:12]
    return predict

def var_P(dT):
    return np.power(1.068, dT)


def var_E(dT):
    return (300 + 25 * (20 + dT) + 0.05 * (20 + dT)**2) / (300 + 25 * 20 + 0.05 * 20**2)


def Q_s(dT):
    eva_sup_season = sarimax_predict(eva_sup)
    eva_hm_season = sarimax_predict(eva_hm)
    eva_eri_season = sarimax_predict(eva_eri)
    eva_ont_season = sarimax_predict(eva_ont)
    prc_sup_season = sarimax_predict(prc_sup)
    prc_hm_season = sarimax_predict(prc_hm)
    prc_eri_season = sarimax_predict(prc_eri)
    prc_ont_season = sarimax_predict(prc_ont)
    runoff_sup_season = sarimax_predict(runoff_sup)
    runoff_hm_season = sarimax_predict(runoff_hur + runoff_mic)
    runoff_eri_season = sarimax_predict(runoff_eri)
    runoff_ont_season = sarimax_predict(runoff_ont)
    Q_s_sup = prc_sup_season * var_P(dT) / 1000 - eva_sup_season * var_E(dT) / 1000 + runoff_sup_season * var_P(dT) * T / area_sup
    Q_s_hm = prc_hm_season * var_P(dT) / 1000 - eva_hm_season * var_E(dT) / 1000 + runoff_hm_season * var_P(dT) * T / area_hm
    Q_s_eri = prc_eri_season * var_P(dT) / 1000 - eva_eri_season * var_E(dT) / 1000 + runoff_eri_season * var_P(dT) * T / area_eri
    Q_s_ont = prc_ont_season * var_P(dT) / 1000 - eva_ont_season * var_E(dT) / 1000 + runoff_ont_season * var_P(dT) * T / area_ont
    Q_s = np.array([Q_s_sup, Q_s_hm, Q_s_eri, Q_s_ont]).T.flatten()
    return Q_s


data = eva_ont
model = SARIMAX(data, order=(1, 0, 0), seasonal_order=(1, 1, 1, 12))
model_fit = model.fit()
predict = model_fit.predict(start=95, end=96 * 2 - 1, dynamic=True)
print(model_fit.summary())


# data = []
# data.append(prc_ont * var_P(0) / 1000)
# data.append(- eva_ont * var_E(0) / 1000)
# data.append(runoff_ont * var_P(0) * T / area_ont)
#
# wb = openpyxl.Workbook()
# sheet = wb.active
# for i in range(3):
#     for j in range(1, 97):
#         sheet.cell(row = j, column = i + 1).value = data[i][j - 1]
# wb.save("output.xlsx")