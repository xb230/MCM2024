import itertools
import openpyxl
import numpy as np
from matplotlib import pyplot as plt
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.api import OLS, add_constant
from pandas.plotting import autocorrelation_plot

'''
-----------------------------------------------------------------------------------------------------------------
Constants
'''

T = 60 * 60 * 24 * 30

area_sup = 81000 * 10 ** 6
area_eri = 25700 * 10 ** 6
area_ont = 19000 * 10 ** 6
area_hm = 118000 * 10 ** 6

'''
-----------------------------------------------------------------------------------------------------------------
Fetch data from excel workbook
'''

wb = openpyxl.load_workbook('data.xlsx', data_only=True)

ws_sup = wb['Lake Superior']
LWL_sup = [ws_sup.cell(row=i, column=j).value
                for i,j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(LWL_sup)):
    LWL_sup[i] = float(LWL_sup[i]) if LWL_sup[i] != '---' else LWL_sup[i-1]
LWL_sup = np.array(LWL_sup)


ws_hm = wb['Lake Michigan and Lake Huron']
LWL_hm = [ws_hm.cell(row=i, column=j).value
                for i,j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(LWL_hm)):
    LWL_hm[i] = float(LWL_hm[i]) if LWL_hm[i] != '---' else LWL_hm[i-1]
LWL_hm = np.array(LWL_hm)


ws_eri = wb['Lake Erie']
LWL_eri = [ws_eri.cell(row=i, column=j).value
            for i,j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(LWL_eri)):
    LWL_eri[i] = float(LWL_eri[i]) if LWL_eri[i] != '---' else LWL_eri[i-1]
LWL_eri = np.array(LWL_eri)


ws_ont = wb['Lake Ontario']
LWL_ont = [ws_ont.cell(row=i, column=j).value
            for i,j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(LWL_ont)):
    LWL_ont[i] = float(LWL_ont[i]) if LWL_ont[i] != '---' else LWL_ont[i-1]
LWL_ont = np.array(LWL_ont)


def get_ont():
    return LWL_ont[60:72]


def get_sup():
    return LWL_sup[0:96]


ws_stMary = wb['St. Mary\'s River']
RFR_stMary = [ws_stMary.cell(row=i, column=j).value
              for i, j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(RFR_stMary)):
    RFR_stMary[i] = float(RFR_stMary[i]) if RFR_stMary[i] != '---' else RFR_stMary[i-1]
RFR_stMary = np.array(RFR_stMary)


ws_det = wb['Detroit River']
RFR_det = [ws_det.cell(row=i, column=j).value
           for i, j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(RFR_det)):
    RFR_det[i] = float(RFR_det[i]) if RFR_det[i] != '---' else RFR_det[i-1]
RFR_det = np.array(RFR_det)


ws_nia = wb['Niagara River']
RFR_nia = [ws_nia.cell(row=i, column=j).value
              for i, j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(RFR_nia)):
    RFR_nia[i] = float(RFR_nia[i]) if RFR_nia[i] != '---' else RFR_nia[i-1]
RFR_nia = np.array(RFR_nia)

ws_law = wb['St. Lawrence River']
RFR_law = [ws_law.cell(row=i, column=j).value
              for i, j in itertools.product(range(20, 31), range(2, 14))]
for i in range(len(RFR_law)):
    RFR_law[i] = float(RFR_law[i]) if RFR_law[i] != '---' else RFR_law[i-1]
RFR_law = np.array(RFR_law)

def flow_law():
    return RFR_law[60:72] * T / area_ont

def flow_mar():
    return RFR_stMary[60:72] * T / area_sup


def h0():
    # h0_sup = np.mean(LWL_sup[60:73])
    # h0_hm = np.mean(LWL_hm[60:73])
    # h0_eri = np.mean(LWL_eri[60:73])
    # h0_ont = np.mean(LWL_ont[60:73])
    # return np.array([h0_sup, h0_hm, h0_eri, h0_ont])
    return np.zeros(4)

def X1():
    x = [-1, area_sup / area_hm, 0, 0]
    R = []
    for i in range(12):
        r = np.array([])
        for j in range(12):
            if i == j:
                r = np.concatenate((r, x))
            else:
                r = np.concatenate((r, [0, 0, 0, 0]))
        R.append(r)
    R = np.array(R).T
    return R


def X2():
    x = [0, 0, 0, -1]
    R = []
    for i in range(12):
        r = np.array([])
        for j in range(12):
            if i == j:
                r = np.concatenate((r, x))
            else:
                r = np.concatenate((r, [0, 0, 0, 0]))
        R.append(r)
    R = np.array(R).T
    return R


def D():
    nia_mean = np.mean(RFR_nia[0:96])
    det_mean = np.mean(RFR_det[0:96])
    D_sup = np.array([0] * 12)
    D_hm = np.array([-det_mean] * 12)
    D_eri = np.array([det_mean - nia_mean] * 12)
    D_ont = np.array([nia_mean] * 12)
    D = np.array([D_sup * T / area_sup, D_hm * T / area_hm, D_eri * T / area_eri, D_ont * T / area_ont]).T.flatten()
    return D

'''
-----------------------------------------------------------------------------------------------------------------
Main
'''



'''
End
-----------------------------------------------------------------------------------------------------------------
'''